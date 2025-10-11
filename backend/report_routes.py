"""
Report Generation Routes
Handles PDF and Excel report generation for water usage data
"""
from fastapi import APIRouter, HTTPException, Depends, Response
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime
from typing import Optional
import os
import io

from auth import get_current_user, require_role
from models import User, UserRole
from analytics_models import ReportRequest, ExportFormat

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference


router = APIRouter(prefix="/reports", tags=["reports"])

# Database connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'indowater_db')]


def parse_date(date_str: str) -> datetime:
    """Parse date string to datetime"""
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except:
        return datetime.strptime(date_str, '%Y-%m-%d')


async def get_usage_data(customer_id: Optional[str], start_date: datetime, end_date: datetime):
    """Fetch usage data for report"""
    query = {
        "reading_date": {
            "$gte": start_date.isoformat(),
            "$lte": end_date.isoformat()
        }
    }
    
    if customer_id:
        query["customer_id"] = customer_id
    
    records = await db.water_usage.find(query, {"_id": 0}).sort("reading_date", 1).to_list(None)
    
    return records


def generate_pdf_report(data: dict, customer_info: dict) -> bytes:
    """Generate PDF report using ReportLab"""
    buffer = io.BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    title = Paragraph("IndoWater Usage Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Report Info
    report_info = [
        ['Report Period:', f"{data['start_date']} to {data['end_date']}"],
        ['Customer:', customer_info.get('full_name', 'All Customers')],
        ['Generated:', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')],
    ]
    
    info_table = Table(report_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1E40AF')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Summary Section
    summary_heading = Paragraph("Usage Summary", heading_style)
    elements.append(summary_heading)
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Consumption', f"{data['total_consumption']:.2f} m³"],
        ['Total Cost', f"Rp {data['total_cost']:,.2f}"],
        ['Average Daily', f"{data['average_daily']:.3f} m³"],
        ['Number of Days', str(data['days'])],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Daily Usage Table (last 30 days)
    if data['records']:
        daily_heading = Paragraph("Daily Usage Details", heading_style)
        elements.append(daily_heading)
        
        # Limit to last 30 records for PDF
        recent_records = data['records'][-30:]
        
        daily_data = [['Date', 'Consumption (m³)', 'Cost (Rp)']]
        for record in recent_records:
            date_str = record['reading_date'][:10]  # Extract date part
            daily_data.append([
                date_str,
                f"{record['consumption']:.3f}",
                f"{record['cost']:,.2f}"
            ])
        
        daily_table = Table(daily_data, colWidths=[2*inch, 2*inch, 2*inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
        ]))
        
        elements.append(daily_table)
    
    # Build PDF
    doc.build(elements)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes


def generate_excel_report(data: dict, customer_info: dict, include_charts: bool = True) -> bytes:
    """Generate Excel report using openpyxl"""
    buffer = io.BytesIO()
    
    # Create workbook
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header
    ws_summary['A1'] = "IndoWater Usage Report"
    ws_summary['A1'].font = Font(size=18, bold=True, color="1E40AF")
    ws_summary.merge_cells('A1:C1')
    
    # Report Info
    ws_summary['A3'] = "Report Period:"
    ws_summary['B3'] = f"{data['start_date']} to {data['end_date']}"
    ws_summary['A4'] = "Customer:"
    ws_summary['B4'] = customer_info.get('full_name', 'All Customers')
    ws_summary['A5'] = "Generated:"
    ws_summary['B5'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    
    # Make labels bold
    for row in range(3, 6):
        ws_summary[f'A{row}'].font = Font(bold=True)
    
    # Summary Data
    ws_summary['A7'] = "Usage Summary"
    ws_summary['A7'].font = Font(size=14, bold=True, color="1E40AF")
    ws_summary.merge_cells('A7:C7')
    
    summary_headers = ['Metric', 'Value']
    ws_summary.append([])
    ws_summary.append(summary_headers)
    
    # Format headers
    header_row = 9
    for col in range(1, 3):
        cell = ws_summary.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Summary values
    summary_rows = [
        ['Total Consumption', f"{data['total_consumption']:.2f} m³"],
        ['Total Cost', f"Rp {data['total_cost']:,.2f}"],
        ['Average Daily', f"{data['average_daily']:.3f} m³"],
        ['Number of Days', str(data['days'])],
    ]
    
    for row in summary_rows:
        ws_summary.append(row)
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    
    # Daily Usage Sheet
    ws_daily = wb.create_sheet("Daily Usage")
    
    # Headers
    ws_daily['A1'] = "Date"
    ws_daily['B1'] = "Consumption (m³)"
    ws_daily['C1'] = "Cost (Rp)"
    ws_daily['D1'] = "Cumulative (m³)"
    
    # Format headers
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_daily[f'{col}1']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    cumulative = 0
    for idx, record in enumerate(data['records'], start=2):
        date_str = record['reading_date'][:10]
        consumption = record['consumption']
        cost = record['cost']
        cumulative += consumption
        
        ws_daily[f'A{idx}'] = date_str
        ws_daily[f'B{idx}'] = round(consumption, 3)
        ws_daily[f'C{idx}'] = round(cost, 2)
        ws_daily[f'D{idx}'] = round(cumulative, 3)
    
    # Adjust column widths
    ws_daily.column_dimensions['A'].width = 15
    ws_daily.column_dimensions['B'].width = 20
    ws_daily.column_dimensions['C'].width = 20
    ws_daily.column_dimensions['D'].width = 20
    
    # Add chart if requested
    if include_charts and len(data['records']) > 1:
        chart = LineChart()
        chart.title = "Daily Water Consumption"
        chart.style = 10
        chart.y_axis.title = 'Consumption (m³)'
        chart.x_axis.title = 'Date'
        
        data_ref = Reference(ws_daily, min_col=2, min_row=1, max_row=len(data['records']) + 1)
        cats_ref = Reference(ws_daily, min_col=1, min_row=2, max_row=len(data['records']) + 1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        ws_daily.add_chart(chart, "F2")
    
    # Save to buffer
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes


@router.post("/export-pdf")
async def export_pdf_report(
    request: ReportRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Generate and download PDF report
    """
    # Permission check
    if current_user.role == UserRole.CUSTOMER:
        customer_id = current_user.id
    else:
        customer_id = request.customer_id
    
    # Parse dates
    start_date = parse_date(request.start_date)
    end_date = parse_date(request.end_date)
    
    # Fetch data
    records = await get_usage_data(customer_id, start_date, end_date)
    
    if not records:
        raise HTTPException(
            status_code=404,
            detail="No usage data found for the specified period"
        )
    
    # Calculate summary
    total_consumption = sum(r['consumption'] for r in records)
    total_cost = sum(r['cost'] for r in records)
    days = (end_date - start_date).days + 1
    average_daily = total_consumption / days if days > 0 else 0
    
    # Get customer info
    if customer_id:
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not customer:
            # Try users collection
            customer = await db.users.find_one({"id": customer_id}, {"_id": 0})
        customer_info = customer if customer else {"full_name": "Unknown"}
    else:
        customer_info = {"full_name": "All Customers"}
    
    # Prepare data
    report_data = {
        "start_date": start_date.strftime('%Y-%m-%d'),
        "end_date": end_date.strftime('%Y-%m-%d'),
        "total_consumption": total_consumption,
        "total_cost": total_cost,
        "average_daily": average_daily,
        "days": days,
        "records": records
    }
    
    # Generate PDF
    pdf_bytes = generate_pdf_report(report_data, customer_info)
    
    # Return as downloadable file
    filename = f"water_usage_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.post("/export-excel")
async def export_excel_report(
    request: ReportRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Generate and download Excel report
    """
    # Permission check
    if current_user.role == UserRole.CUSTOMER:
        customer_id = current_user.id
    else:
        customer_id = request.customer_id
    
    # Parse dates
    start_date = parse_date(request.start_date)
    end_date = parse_date(request.end_date)
    
    # Fetch data
    records = await get_usage_data(customer_id, start_date, end_date)
    
    if not records:
        raise HTTPException(
            status_code=404,
            detail="No usage data found for the specified period"
        )
    
    # Calculate summary
    total_consumption = sum(r['consumption'] for r in records)
    total_cost = sum(r['cost'] for r in records)
    days = (end_date - start_date).days + 1
    average_daily = total_consumption / days if days > 0 else 0
    
    # Get customer info
    if customer_id:
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not customer:
            # Try users collection
            customer = await db.users.find_one({"id": customer_id}, {"_id": 0})
        customer_info = customer if customer else {"full_name": "Unknown"}
    else:
        customer_info = {"full_name": "All Customers"}
    
    # Prepare data
    report_data = {
        "start_date": start_date.strftime('%Y-%m-%d'),
        "end_date": end_date.strftime('%Y-%m-%d'),
        "total_consumption": total_consumption,
        "total_cost": total_cost,
        "average_daily": average_daily,
        "days": days,
        "records": records
    }
    
    # Generate Excel
    excel_bytes = generate_excel_report(report_data, customer_info, request.include_charts)
    
    # Return as downloadable file
    filename = f"water_usage_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
